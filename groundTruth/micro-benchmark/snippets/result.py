import os
import json
from dotenv import load_dotenv

load_dotenv()


def run(truth, pycg, pythoncg):
    if isinstance(truth, int) or isinstance(pycg, int) or isinstance(pycg, int):
        return
    TP, FN, FP = 0, 0, 0
    with open(truth, "r") as f:
        truthJson: dict = json.load(f)
    with open(pycg, "r") as f:
        pycgJson: dict = json.load(f)
    with open(pythoncg, "r") as f:
        pythoncgJson: dict = json.load(f)

    # Result array contains [TP, FP, FN, edges]
    pycg_res = getResult(truthJson, pycgJson)
    python_res = getResult(truthJson, pythoncgJson)

    return pycg_res, python_res


def getEdges(cg: dict):
    cnt = 0
    for k, vList in cg.items():
        cnt += len(vList)
    return cnt


def getTP(truthJson, curJson):
    cnt = 0
    for k, vList in curJson.items():
        if k in truthJson:
            cnt += len(set(vList) & set(truthJson[k]))
    return cnt


def getFN(truthJson, curJson, TP):
    return getEdges(truthJson) - TP


def getFP(truthJson, curJson, TP):
    return getEdges(curJson) - TP


def getResult(truthJson: dict, curJson: dict):
    TP = getTP(truthJson, curJson)
    FP = getFP(truthJson, curJson, TP)
    FN = getFN(truthJson, curJson, TP)
    edges = getEdges(truthJson)
    return [TP, FP, FN, edges]


def getMem(pycgPath, pythonPath):
    pycgmem = 0
    pythonmem = 0
    with open(pycgPath, "r") as f:
        lines = f.read().split("\n")
    for line in lines:
        line = line.strip()
        if line.endswith("maximum resident set size"):
            res = line.replace("maximum resident set size", "")
            res = int(res)
            pycgmem = res
    with open(pythonPath, "r") as f:
        lines = f.read().split("\n")
    for line in lines:
        line = line.strip()
        if line.endswith("maximum resident set size"):
            res = line.replace("maximum resident set size", "")
            res = int(res)
            pythonmem = res
            break
    return pycgmem, pythonmem


def findFile(base):
    for root, ds, fs in os.walk(base):
        returnList = [0] * 3
        for f in fs:
            if f == "callgraph.json":
                fullname = os.path.join(root, f)
                returnList[0] = fullname
            if f == "pycg.json":
                fullname = os.path.join(root, f)
                returnList[1] = fullname
            if f == "pythonCG.json":
                fullname = os.path.join(root, f)
                returnList[2] = fullname
        yield run(returnList[0], returnList[1], returnList[2])
        # yield getMem(returnList[0], returnList[1])


global_pycg = [0, 0, 0, 0]
global_pythoncg = [0, 0, 0, 0]


def main(index, base):
    def save_xlsx(index, name, res):
        from openpyxl import load_workbook

        filename = os.environ.get("SNIPPETS_PATH") + "/micro.xlsx"
        wb = load_workbook(filename=filename)
        sheet = wb["Sheet1"]

        sheet.cell(1, 3, "PyCG")
        sheet.cell(1, 8, "Jarvis")

        for i, value in enumerate(["TP", "FP", "FN", "edges", ""] * 2):
            sheet.cell(2, i + 3, value)

        tmprow = index
        sheet.cell(tmprow, 1, value=name)
        col = 3
        for j, tmp in enumerate(res[:4]):
            tmpcol = col + j
            sheet.cell(tmprow, tmpcol, value=tmp)
        col = 8
        for j, tmp in enumerate(res[4:]):
            tmpcol = col + j
            sheet.cell(tmprow, tmpcol, value=tmp)
        print()
        wb.save(filename)

    global global_pycg
    global global_pythoncg
    pre_pycg = [0, 0, 0, 0]
    pre_python = [0, 0, 0, 0]
    for i in findFile(base):
        if not i:
            continue
        pre_pycg = list(map(lambda x: x[0] + x[1], zip(pre_pycg, list(i[0]))))
        pre_python = list(map(lambda x: x[0] + x[1], zip(pre_python, list(i[1]))))

    pycg_str = "{},{},{}/{}".format(*pre_pycg)
    python_str = "{},{},{}/{}".format(*pre_python)
    res = pre_pycg + pre_python
    save_xlsx(3 + index, base.split(os.path.sep)[-1], res)
    print(base.split(os.path.sep)[-1])
    print(pycg_str, python_str)
    global_pycg = list(map(lambda x: x[0] + x[1], zip(pre_pycg, global_pycg)))
    global_pythoncg = list(map(lambda x: x[0] + x[1], zip(pre_python, global_pythoncg)))


SNIPPETS_PATH = os.environ.get("SNIPPETS_PATH")
entries = [
    f"{SNIPPETS_PATH}/assignments",
    f"{SNIPPETS_PATH}/builtins",
    f"{SNIPPETS_PATH}/classes",
    f"{SNIPPETS_PATH}/decorators",
    f"{SNIPPETS_PATH}/dicts",
    f"{SNIPPETS_PATH}/direct_calls",
    f"{SNIPPETS_PATH}/exceptions",
    f"{SNIPPETS_PATH}/functions",
    f"{SNIPPETS_PATH}/generators",
    f"{SNIPPETS_PATH}/imports",
    f"{SNIPPETS_PATH}/kwargs",
    f"{SNIPPETS_PATH}/lambdas",
    f"{SNIPPETS_PATH}/lists",
    f"{SNIPPETS_PATH}/mro",
    f"{SNIPPETS_PATH}/args",
    f"{SNIPPETS_PATH}/returns",
    # New cases
    f"{SNIPPETS_PATH}/newCase/args",
    f"{SNIPPETS_PATH}/newCase/assign",
    f"{SNIPPETS_PATH}/newCase/calls",
    f"{SNIPPETS_PATH}/newCase/control_flow",
    f"{SNIPPETS_PATH}/newCase/import",
]

if __name__ == "__main__":
    for index, entry in enumerate(entries):
        main(index, entry)
    print(global_pycg, global_pythoncg)

# complete 不包含没有调用的边
# sound    包含了所有调用的边
